#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""把排障步骤表（xlsx）改写为 markdown skill。

和主蒸馏流水线的区别在输入：主蒸馏吃的是散文式排障文档，要靠模型把判据提炼出来；
这里的输入已经是判据化的表格，模型要做的是**改写而不是提炼**——表里的"步骤详细描述"
习惯把多个分支塞进一长句：

    ……如果不存在则返回"X不存在"并结束执行；如果存在，则判断Policy State是否为Up，
    如果不是Up，则继续向下执行；如果……则返回"状态正常"并结束执行。

一句话里含 4 个分支、2 个终止点，agent 读这种句子容易漏分支。改写后每个分支单独一行、
写清判据字段与去向。命令也要统一：表里 `<endpointipv6>`、`<colorid>`、
`<segment-list-id>` 三种参数写法混用，输出里统一成 <小写-连字符> 并用反引号包起来。

流程：
  1. 解析 xlsx：一个 sheet 多个场景，靠首列的**合并单元格**分块（一个合并区间 =
     一个场景，区间覆盖的行是它的排障步骤）；
  2. 体检：ragIndex 是命令的稳定编号，重号会让命令查不准，解析阶段就报出来；
  3. 改写：每个场景一次模型调用，产出一篇完整 skill；
  4. 校验 + 报告：格式不合规的不落盘，只记进报告。

用法：
  python3 excel_skill_distill_pipeline.py --check    # 只解析和体检，不调模型
  python3 excel_skill_distill_pipeline.py            # 按文件末尾 main() 的默认参数运行
"""
import json
import os
import re
import sys
from datetime import datetime
from multiprocessing import Pool

from skill_self_distill_pipeline import call_model_with_retry, extract_markdown_content
from skill_case_merge_pipeline import BANNED_CONTENT_PATTERNS, extract_json_block

# 表头所在行，其余行为数据行
HEADER_ROW = 1

# 列序（1基），跟着表头走。改表结构时只动这里。
COL_GOAL = 1        # 排障目标：告警名 + 故障构造方法
COL_TOPOLOGY = 2    # 组网场景
COL_STEP_NO = 3     # 排障步骤编号
COL_STEP_DESC = 4   # 排障步骤描述
COL_STEP_DETAIL = 5  # 步骤详细描述（含分支逻辑，改写的主要对象）
COL_RAG_INDEX = 6   # 命令行编号及用途（ragIndex）
COL_COMMAND = 7     # 命令行
COL_ECHO = 8        # 回显
COL_FIX = 9         # 配置修复建议
COL_IMPACT = 10     # 修复建议影响性
COL_VERIFY = 11     # 修复验证

# 场景名 → skill 相对路径。不写在这里时按 derive_skill_path 从告警名推。
# 让模型自己编路径会导致同一张表重跑生成不同文件名（案例合并流水线上已经吃过这个亏，
# 见 skill_case_merge_pipeline.TARGET_OVERRIDES），所以路径由脚本定、不交给模型。
SCENARIO_PATH_OVERRIDES = {}

# 默认的一级目录：这批 skill 独立于既有 skill 库，不并进"故障处理："那套目录
DEFAULT_CATEGORY = "排障步骤"


# 这批 skill 的写作约束单独一份，不复用 skill_case_merge_pipeline.WRITING_RULES。
# 那份是三条现役流水线在吃的，而这里的输入有两处和它直接冲突：
#   1. 那份禁止把"仿真验证"写成步骤（agent 执行不了），但本表的"修复建议影响性"列
#      明写了影响大时要靠仿真做可靠性保障——这属于交给人的风险提示，不是 agent 的步骤，
#      所以这里改成写进"影响性"而不是写成排查步骤；
#   2. 那份要求不要照搬回显，但本表的判据直接依赖回显里的确切字段名
#      （Policy State / List State / Verification State），字段名必须留下。
WRITING_RULES = """- 输出面向网管agent执行，凡是收集信息、联系技术支持、提交给工程师这类动作，整个步骤删除。
- 输出全文禁止出现"联系技术支持"、"寻求技术支持"、"提交给工程师"、"收集信息并联系"等表述，排障步骤穷尽后直接结束。
- skill的读者是执行排障的agent，它看不到这张表。正文里禁止出现"按表格描述"、"表中未给出"、"本步骤"、"原表"这类交代来源的话，也不要写"1号命令行"、"执行2号命令"这种表内编号——要把编号换成真正的命令。
- 回显不要整段照搬，但**判据依赖的字段名必须原样保留并用反引号标出**（如 `Policy State`、`List State`、`Verification State`、`BFD State`），agent 要靠这些字段名在回显里定位。字段的取值同样照写（如 `Down (Shutdown)`、`Down (Overrun)`、`SID Unreachable`）。
- 表里"修复建议影响性"列的内容（如影响大需要仿真验证）写进该修复方案的影响性说明，**不要写成排查步骤**——仿真是交给人的风险提示，agent 执行不了。
- 表里"修复验证"列的内容写成该修复方案之后的验证动作，给出验证命令和期望看到的状态。
- 只写表里给出的信息。表没给的 HTTP 方法、参数、命令一律不要自己编。表里只写了"减少policy数量"这种没有具体命令的修复方向时，就照实写方向，**不要编出一条 CLI 来**；也不要编造表里没有的查询命令（如 `display xxx summary`）来做验证。
- **修复CLI里禁止出现从回显样例抄来的具体值**。回显里的 `bgp 100`、`segment-list 1`、`policy1`、`1::1` 都是某台设备当时的取值，换一台就是错的：AS号、policy名、segment-list名、接口名、IP一律写成 `<as-number>`、`<policy-name>`、`<segment-list-name>` 这样的参数。回显只用来说明"该看哪个字段"，不是配置模板。
- 修复手段和复检命令**只写在根因对照表里**，排查步骤的「根因定位」只给根因名称——同一份修复在两处各写一遍，改了一处忘另一处就会互相矛盾。"""


# 输出格式的定义放在单独的模板文件里，流水线运行时读进来拼进 prompt。
# 这样改格式要求只要改那个 markdown，不用动代码——模板本来就是给人看、给人改的。
SKILL_TEMPLATE_PATH = os.path.join(
    os.path.dirname(os.path.abspath(__file__)), "skill_template.md")

TEMPLATE_BODY = re.compile(r"<template>\s*(.+?)\s*</template>", re.DOTALL)


def load_skill_template(path: str = SKILL_TEMPLATE_PATH) -> str:
    """读模板文件里 <template>…</template> 之间的正文。

    标签外面是给维护者看的说明（这文件是干嘛的、改完怎么验），不该进 prompt。
    """
    if not os.path.isfile(path):
        raise FileNotFoundError(
            f"找不到skill模板 {path}；它定义了生成的skill长什么样，是prompt的一部分")
    text = open(path, encoding="utf-8").read()
    match = TEMPLATE_BODY.search(text)
    if not match:
        raise ValueError(
            f"{path} 里没有 <template>…</template> 标记，无法确定哪一段是模板正文")
    body = match.group(1).strip()
    if not body:
        raise ValueError(f"{path} 的 <template> 里是空的")
    return body


# 这两段不能进模板文件：它们要按场景现填（参数清单、根因清单从步骤表里抽），
# 而模板是静态的、对所有场景通用。
SCENARIO_SPEC = """
# 本场景的入参

本场景的命令已经用到下列参数，入参列表必须覆盖它们：<derived_params>

# 本场景的根因清单（必须逐字使用）

输入表格的"步骤详细描述"里已经用 提示/返回"xxx" 的形式写明了每一步的判定结论，本场景的根因就是下面这些：

<root_causes>

**这些名称要原样使用**，不要改写、简化或另起名字（`{endpoint/color}` 这类占位可以去掉或替换成实际参数）。排查步骤的「根因定位」和根因对照表的「根因」列都用它们，两处必须一致。清单之外不要自己发明根因；清单里的每一条都必须有对应的排查步骤把它判出来。

清单里标了「无故障分支」的那条（如"……状态正常"）**也要在流程里有明确落点**：写清依据哪些字段取值判定为正常、然后结束排查，并在根因对照表里占一行（「修复CLI和方法」写"无需修复"）。agent 排到这一支时需要有对照可依，漏掉它会让 agent 以为自己少判了一步。

# 分支必须逐条展开（改写的核心）

输入表格的"步骤详细描述"习惯把多个分支塞进一长句，例如：

    判断X是否存在，如果不存在则返回"X不存在"并结束执行；如果存在，则判断Policy State是否为Up，
    如果不是Up，则继续向下执行；如果是Up，则判断List State是否为Up，如果不是，则继续向下执行；
    如果都是Up，则返回"状态正常"并结束执行。

这种写法agent容易漏分支，必须按模板的格式拆开：每个分支单独一行、写清判据字段与取值、写清去向（跳转步骤N / 结束排查）。不要写"继续向下执行"这种没有具体去向的说法。

不要保留输入表里的"1号命令行"、"执行4号命令"这类编号说法，直接写出命令本身——agent 看不到那张表。
"""


PROMPT_TEMPLATE = """你是IPRAN网络运维专家，需要把一份排障步骤表改写成供网管agent使用的skill。

表里的内容已经是判据化的（每一步都写明了查什么、看哪个字段、什么取值走哪个分支），你的任务**不是重新提炼判据**，而是把它改写成结构清晰、分支明确的skill：把塞在一长句里的多个分支拆成逐条，把表内的命令编号换成真正的命令，把参数写法统一。判据本身要忠实于表格，不要自己改判断条件。

# 输入：排障步骤表中的一个故障场景
<scenario>

# 输出格式要求
<format_spec>

# 内容要求
<writing_rules>
- 输出完整的skill，以frontmatter开头，frontmatter 之后直接按模板写四个章节（从一级标题 `# 入参列表` 开始），不要再另加篇名标题。frontmatter 的 name 用英文小写+连字符；description 写「故障现象 + 适用时机」，例如"SRv6 TE Policy Down（隧道中断）。出现 SRv6 隧道不通 / SR-Policy 状态异常等告警时使用，覆盖配置缺失、BFD Down、规格超限等场景。"
- 步骤表里靠同一条命令的回显区分的多个根因，把那条命令放进前置检查采集一次，各步骤复用它的回显——不要每步重复下发。据此重新编排步骤是允许的，但判据本身要忠实于表格。
- 表里"修复建议影响性"列的内容写进对应根因的「修复」里作为影响性提示；"修复验证"列的内容写成该根因的「复检命令」。

# 输出格式
先输出一个json代码块给出改写说明，再输出一个markdown代码块给出skill全文：
```json
{"scenario": "故障场景名", "steps": 步骤数, "branches_expanded": "把哪几步的长句拆成了逐条分支", "commands_normalized": [{"from": "<endpointipv6>", "to": "<endpoint-ipv6>"}]}
```
```markdown
（skill全文）
```
"""


def _cell(ws, row: int, col: int) -> str:
    """读单元格并规整为字符串；合并区间的非锚点格在 openpyxl 里读出来是 None。"""
    value = ws.cell(row=row, column=col).value
    if value is None:
        return ""
    return str(value).strip()


RAG_INDEX_PATTERN = re.compile(r"^\s*(\d+)\s*[：:、.,]\s*(.*)$")


def parse_rag_index(raw: str) -> tuple:
    """把"1：查询SRv6 TE Policy"拆成 (编号, 用途)。

    分隔符在表里就不统一（全角冒号、半角冒号、顿号都有），所以几种都认。
    拆不出编号时返回 (None, 原文)，交给体检环节报出来。
    """
    match = RAG_INDEX_PATTERN.match(raw or "")
    if not match:
        return None, (raw or "").strip()
    return int(match.group(1)), match.group(2).strip()


def parse_sheet(xlsx_path: str, sheet_name: str = None) -> list:
    """解析排障步骤表，返回场景列表。

    一个 sheet 放多个场景，靠首列分块：首列非空的行是一个场景的开始，到下一个
    首列非空的行之前都属于这个场景。合并单元格天然满足这个规则（openpyxl 里
    合并区间只有左上角那一格有值），单步场景没有合并区间也照样能切。
    """
    from openpyxl import load_workbook

    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb[wb.sheetnames[0]]

    starts = [row for row in range(HEADER_ROW + 1, ws.max_row + 1)
              if _cell(ws, row, COL_GOAL)]
    if not starts:
        raise ValueError(
            f"{xlsx_path} 的 sheet {ws.title!r} 里没有解析到任何场景："
            f"首列（排障目标）全为空。一个场景的所有步骤行要用合并单元格圈起来，"
            f"或至少在该场景第一行的首列写上排障目标。")

    scenarios = []
    for i, start in enumerate(starts):
        end = starts[i + 1] - 1 if i + 1 < len(starts) else ws.max_row
        steps = []
        for row in range(start, end + 1):
            # 整行都空的尾行（Excel 常留空行）跳过
            if not any(_cell(ws, row, col) for col in
                       (COL_STEP_NO, COL_STEP_DESC, COL_STEP_DETAIL, COL_COMMAND)):
                continue
            rag_no, rag_purpose = parse_rag_index(_cell(ws, row, COL_RAG_INDEX))
            steps.append({
                "row": row,
                "no": _cell(ws, row, COL_STEP_NO),
                "desc": _cell(ws, row, COL_STEP_DESC),
                "detail": _cell(ws, row, COL_STEP_DETAIL),
                "rag_no": rag_no,
                "rag_purpose": rag_purpose,
                "rag_raw": _cell(ws, row, COL_RAG_INDEX),
                "command": _cell(ws, row, COL_COMMAND),
                "echo": _cell(ws, row, COL_ECHO),
                "fix": _cell(ws, row, COL_FIX),
                "impact": _cell(ws, row, COL_IMPACT),
                "verify": _cell(ws, row, COL_VERIFY),
            })
        goal = _cell(ws, start, COL_GOAL)
        scenarios.append({
            "goal": goal,
            "name": goal.splitlines()[0].strip() if goal else f"场景{i + 1}",
            "topology": _cell(ws, start, COL_TOPOLOGY),
            "steps": steps,
            "rows": (start, end),
            "sheet": ws.title,
        })
    return scenarios


def audit_commands(scenario: dict) -> list:
    """体检 ragIndex：它是命令的稳定编号（给RAG检索用），重号会让命令查不准。

    返回问题说明列表（空列表表示没问题）。这类问题在解析阶段就该报出来，
    不能带进 skill——步骤正文里写着"执行6号命令"，而6号对应两条不同的命令时，
    改写出来的skill就会指向错的那条。
    """
    issues = []
    by_index = {}
    for step in scenario["steps"]:
        if step["rag_no"] is None:
            if step["rag_raw"]:
                issues.append(
                    f"步骤{step['no']}（第{step['row']}行）的ragIndex "
                    f"{step['rag_raw']!r} 解析不出编号，需写成"
                    f"“编号: 用途”的形式")
            continue
        by_index.setdefault(step["rag_no"], []).append(step)

    for index, steps in sorted(by_index.items()):
        commands = {s["command"] for s in steps if s["command"]}
        if len(commands) > 1:
            detail = "；".join(
                f"步骤{s['no']}用 {s['command']!r}" for s in steps if s["command"])
            issues.append(
                f"ragIndex {index} 被 {len(steps)} 个步骤用在了 {len(commands)} 条"
                f"不同的命令上：{detail}。ragIndex 是命令的稳定编号，必须一号一命令")
    return issues


def audit_step_numbers(scenario: dict) -> list:
    """步骤编号应为从1开始的连续整数——正文里的"继续执行步骤N"依赖它。"""
    issues = []
    numbers = []
    for step in scenario["steps"]:
        try:
            numbers.append(int(float(step["no"])))
        except (TypeError, ValueError):
            issues.append(f"第{step['row']}行的步骤编号 {step['no']!r} 不是整数")
    if numbers and numbers != list(range(1, len(numbers) + 1)):
        issues.append(f"步骤编号不是从1开始的连续整数: {numbers}")
    return issues


# 表里"步骤详细描述"用 提示/返回"xxx" 的形式明确写出了每一步的判定结论，
# 根因名称就在里面。抽出来当权威清单交给模型，比让它自己起名字可靠得多——
# 名字一旦被改写，根因对照表就和排查步骤对不上了。
CONCLUSION_PATTERN = re.compile(
    r"(?:返回|提示)(?:错误信息)?\s*[“\"]([^”\"]+)[”\"]")


def _normalize_cause(text: str) -> str:
    """比对根因名用：去掉 {endpoint/color} 这类占位、空白，并转小写。"""
    return re.sub(r"\s+", "", re.sub(r"[{（(][^}）)]*[}）)]", "", text or "")).lower()


def extract_root_causes(scenario: dict) -> list:
    """从步骤详细描述里抽出每一步的判定结论。

    "……状态正常"这类**无故障分支**同样算一条结论，要一并写进排查步骤和根因对照表：
    agent 排到那一支时也需要有对照可依，否则会以为漏了判断。只是它的「修复」是
    "无需修复"。normal 字段用来在 prompt 里标注这一点。
    """
    causes = []
    seen = set()
    for step in scenario["steps"]:
        for text in CONCLUSION_PATTERN.findall(step["detail"] or ""):
            key = _normalize_cause(text)
            if key and key not in seen:
                seen.add(key)
                causes.append({"step": step["no"], "text": text.strip(),
                               "normal": "正常" in text})
    return causes


def collect_parameters(scenario: dict) -> list:
    """表里命令用到的全部 <参数>，按出现顺序去重。"""
    params, seen = [], set()
    for step in scenario["steps"]:
        for field in ("command", "fix", "verify"):
            for param in re.findall(r"<([^>\n]+)>", step.get(field) or ""):
                if param not in seen:
                    seen.add(param)
                    params.append(param)
    return params


def derive_skill_path(scenario: dict, category: str = DEFAULT_CATEGORY) -> str:
    """从告警名推出 skill 相对路径。

    交给模型编路径会让同一张表重跑生成不同文件名，所以这里用确定性规则定死。
    """
    if scenario["name"] in SCENARIO_PATH_OVERRIDES:
        return SCENARIO_PATH_OVERRIDES[scenario["name"]]
    base = re.sub(r"[/\\:*?\"<>|]", "_", scenario["name"]).strip()
    base = re.sub(r"(告警|事件)$", "", base).strip() or "未命名场景"
    return f"{category}/{base}.md"


def format_scenario(scenario: dict) -> str:
    """把一个场景渲染成喂给模型的文本。

    命令与回显按 ragIndex 单独列一份"命令清单"：表里同一条命令会被多个步骤引用
    （步骤4~8 都在读步骤1 的回显），清单能让模型看清哪些步骤共用一次命令执行，
    避免改写成每个分支各跑一遍同样的命令。
    """
    lines = [f"## 排障目标\n{scenario['goal']}"]
    if scenario["topology"]:
        lines.append(f"\n## 组网场景\n{scenario['topology']}")

    # 清单按**命令本身**去重，不按 ragIndex：表里 ragIndex 可能重号（同一个号
    # 指向两条不同的命令），按号去重会让后一条被静默丢掉，清单于是和步骤里写的
    # 命令对不上——模型拿到自相矛盾的映射，多半会给那一步写错命令。
    commands = {}
    for step in scenario["steps"]:
        if not step["command"]:
            continue
        entry = commands.setdefault(step["command"], {
            "indexes": [], "purpose": step["rag_purpose"], "echo": step["echo"],
            "first_row": step["row"]})
        if step["rag_no"] is not None and step["rag_no"] not in entry["indexes"]:
            entry["indexes"].append(step["rag_no"])
        if not entry["echo"] and step["echo"]:
            entry["echo"] = step["echo"]

    # 一个号指向多条命令时，明确告诉模型别按号认命令，按步骤自己写的命令行认
    by_index = {}
    for command, entry in commands.items():
        for index in entry["indexes"]:
            by_index.setdefault(index, []).append(command)
    ambiguous = sorted(i for i, cmds in by_index.items() if len(cmds) > 1)

    if commands:
        block = ["\n## 命令清单（同一条命令只执行一次，多个步骤共用它的回显）"]
        if ambiguous:
            block.append(
                f"注意：编号 {'、'.join(str(i) for i in ambiguous)} 在本表里被用在了"
                f"多条不同的命令上。步骤正文里出现的“N号命令”因此不可靠，"
                f"请一律以该步骤自己列出的“使用命令”那一行为准。")
        for command, entry in sorted(commands.items(),
                                     key=lambda kv: kv[1]["first_row"]):
            label = "、".join(f"{i}号" for i in entry["indexes"]) or "未编号"
            block.append(f"\n### {label}：{entry['purpose']}\n命令行：{command}")
            if entry["echo"]:
                block.append(f"回显：\n```\n{entry['echo']}\n```")
        lines.append("\n".join(block))

    lines.append("\n## 排障步骤")
    for step in scenario["steps"]:
        block = [f"\n### 步骤{step['no']}：{step['desc']}",
                 f"详细描述：{step['detail']}"]
        if step["rag_no"] is not None and step["command"]:
            block.append(f"使用命令{step['rag_no']}（{step['rag_purpose']}）："
                         f"{step['command']}")
        elif step["command"]:
            block.append(f"使用命令：{step['command']}")
        else:
            block.append("本步骤不执行新命令，读前面某条命令的回显即可")
        if step["fix"]:
            block.append(f"配置修复建议：{step['fix']}")
        if step["impact"]:
            block.append(f"修复建议影响性：{step['impact']}")
        if step["verify"]:
            block.append(f"修复验证：{step['verify']}")
        lines.append("\n".join(block))
    return "\n".join(lines)


def build_format_spec(scenario: dict) -> str:
    """模板文件 + 本场景抽出来的参数与根因，拼成 prompt 的格式要求部分。

    参数和根因都能从表里确定性地拿到，交给模型自己想只会想歪：参数名会写混，
    根因名会被改写成同义词，改写完根因对照表就和排查步骤对不上了。
    """
    params = collect_parameters(scenario)
    causes = extract_root_causes(scenario)
    return (load_skill_template() + "\n\n" + SCENARIO_SPEC
            .replace("<derived_params>",
                     "、".join(f"`<{p}>`" for p in params) or "（表里的命令没有参数）")
            .replace("<root_causes>",
                     "\n".join(
                         f"{i}. {c['text']}（来自步骤{c['step']}"
                         f"{'，无故障分支' if c['normal'] else ''}）"
                         for i, c in enumerate(causes, 1))
                     or "（表里没有明确写出判定结论，按步骤描述自行归纳）"))


# ---------------------------------------------------------------- 输出校验

REQUIRED_SECTIONS = ["入参列表", "前置检查", "排查步骤", "根因对照表"]

# 步骤标题：`### 1. 判定静态配置完整性`，也认 `### 步骤 1：xxx` 这种写法。
# 前置检查用的是同样的编号形式，靠按 `## ` 小节切分来区分，不靠标题本身。
STEP_HEADING = re.compile(
    r"^#{2,4}\s*(?:步骤\s*)?(\d+)\s*[.、：:]\s*(.+)$", re.MULTILINE)
# 跳转引用：跳转步骤 5 / 转步骤5 / 进入排查步骤 1
STEP_REFERENCE = re.compile(r"(?:跳转|转入?|进入)\s*(?:排查)?步骤\s*(\d+)")


# 只认这四个章节名作切分点，且 # / ## 两级都收。
# 模板要求章节用一级标题（`# 入参列表`）、步骤用二级（`## 1. 名称`）；但按标题级别
# 切会把步骤标题也当成章节，按名字切才稳，顺带兼容写成 `## 入参列表` 的产出。
SECTION_HEADING = re.compile(
    r"^#{1,2}\s*(" + "|".join(REQUIRED_SECTIONS) + r")\s*$", re.MULTILINE)


def split_sections(content: str) -> dict:
    """按四个章节标题切分正文，返回 {章节名: 正文}。"""
    sections = {}
    matches = list(SECTION_HEADING.finditer(content))
    for i, match in enumerate(matches):
        end = matches[i + 1].start() if i + 1 < len(matches) else len(content)
        sections[match.group(1)] = content[match.end():end]
    return sections


def table_first_column(section: str) -> list:
    """取markdown表格第一列的值（跳过表头与分隔行）。"""
    values = []
    for line in section.splitlines():
        line = line.strip()
        if not line.startswith("|"):
            continue
        cells = [c.strip() for c in line.strip("|").split("|")]
        if not cells or not cells[0] or set(cells[0]) <= set("-: "):
            continue
        values.append(cells[0])
    return values[1:] if values else []      # 首行是表头


def check_structure(content: str) -> str:
    """四个小节齐不齐、顺序对不对。"""
    sections = split_sections(content)
    missing = [name for name in REQUIRED_SECTIONS if name not in sections]
    if missing:
        return (f"缺少必需的小节: {'、'.join('## ' + m for m in missing)}"
                f"（现有小节: {'、'.join(sections) or '无'}）")
    order = [name for name in sections if name in REQUIRED_SECTIONS]
    if order != REQUIRED_SECTIONS:
        return f"小节顺序不对: 应为 {' → '.join(REQUIRED_SECTIONS)}，实际为 {' → '.join(order)}"
    return ""


# 前置检查里的真跳转。只认"跳转步骤N"：
#   - "跳转信息"是字段名，跳转后面不跟步骤号，不算；
#   - "继续前置检查步骤2"是线性推进到下一条，是允许的；
#   - "进入排查步骤1"是从前置检查走进排查步骤，也是允许的。
PRECHECK_JUMP = re.compile(r"跳转\s*(?:到\s*)?步骤\s*(\d+)")


def check_precheck_linear(content: str) -> str:
    """前置检查里不许有内部跳转。"""
    section = split_sections(content).get("前置检查", "")
    hit = PRECHECK_JUMP.search(section)
    if hit:
        return (f"前置检查里出现了跳转“{hit.group(0)}”——前置检查必须是线性执行的"
                f"信息采集，分支判断放到排查步骤里")
    return ""


def table_rows(section: str) -> list:
    """取markdown表格的数据行（去掉表头与分隔行），每行为单元格列表。"""
    rows = []
    for line in section.splitlines():
        line = line.strip()
        if not line.startswith("|"):
            continue
        cells = [c.strip() for c in line.strip("|").split("|")]
        if not cells or not cells[0] or set(cells[0]) <= set("-: "):
            continue
        rows.append(cells)
    return rows[1:] if rows else []


def check_declared_params(content: str) -> str:
    """前置检查用到的参数必须是入参列表里的必填项。

    只约束前置检查：排查步骤里的参数（如从回显才能拿到的 <segment-list-id>）
    本来就允许是选填的。
    """
    sections = split_sections(content)
    rows = table_rows(sections.get("入参列表", ""))
    if not rows:
        return "入参列表里没有解析到任何行，需要是三列表格（信息/是否必填/说明）"
    required = {re.sub(r"[\s\-_]", "", row[0]).lower()
                for row in rows if len(row) > 1 and row[1].strip() in ("是", "Y", "yes")}
    if not required:
        return "入参列表里没有任何必填项（「是否必填」列为「是」的行）"
    for span in INLINE_CODE.findall(sections.get("前置检查", "")):
        for param in re.findall(r"<([^>\n]+)>", span):
            key = re.sub(r"[\s\-_]", "", param).lower()
            if not any(key == r or key in r or r in key for r in required):
                return (f"前置检查的命令用了参数 `<{param}>`，但入参列表里没有对应行——"
                        f"前置检查的参数不可以超出入参列表")
    return ""


def check_root_causes(content: str, scenario: dict) -> str:
    """根因对照表要覆盖步骤表里写明的每一个根因，正文里也得真的判到。

    以**根因对照表**为准，而不是去正文里找某种固定标记：正文里根因的写法本来就
    多样（`根因定位：X`、`→ "X"`、`根因：**X**`），锁死一种写法只会逼模型反复重试。
    对照表是结构化的，从它取根因可靠得多。
    """
    sections = split_sections(content)
    in_table = {_normalize_cause(c): c
                for c in table_first_column(sections.get("根因对照表", ""))}
    if not in_table:
        return ("根因对照表里没有解析到任何行，需要是四列表格"
                "（根因/现象/修复CLI和方法/复检命令）")

    body = sections.get("前置检查", "") + "\n" + sections.get("排查步骤", "")
    body_key = _normalize_cause(body)

    faults = [c for c in extract_root_causes(scenario) if not c["normal"]]
    absent_from_table = [c["text"] for c in faults
                         if _normalize_cause(c["text"]) not in in_table]
    if absent_from_table:
        return (f"根因对照表漏了步骤表里写明的根因: {'、'.join(absent_from_table)}"
                f"——对照表必须覆盖每一个根因")

    absent_from_body = [c["text"] for c in faults
                        if _normalize_cause(c["text"]) not in body_key]
    if absent_from_body:
        return (f"这些根因只在对照表里有、正文里没有判到: {'、'.join(absent_from_body)}"
                f"——每个根因都要有对应的排查步骤把它判出来")

    # 无故障分支（"……状态正常"）要有落点，否则agent排到这一支会以为自己漏判了。
    # 措辞不强求和步骤表逐字一致（正文里常写成"隧道状态正常"），只要判到了即可。
    if any(c["normal"] for c in extract_root_causes(scenario)) and "正常" not in body:
        return ("正文里没有无故障分支的判定（步骤表写明了“……状态正常”这一支）"
                "——排到这一支时agent需要有明确结论，否则会以为自己漏判了")
    return ""



INLINE_CODE = re.compile(r"`([^`\n]+)`")
# 命令里残留的非 <> 占位符：{xxx}、[xxx]、连续大写占位
BAD_PLACEHOLDER = re.compile(r"\{[^}\n]+\}|\[[a-z][^\]\n]*\]|\bXXX+\b")
# 表内编号说法，改写后不该留下
TABLE_REFERENCE = re.compile(r"\d+\s*号命令(行)?|执行\s*\d+\s*号")


FENCE_BLOCK = re.compile(r"```[^\n]*\n(.*?)```", re.DOTALL)
QUERY_PREFIX = re.compile(r"^(display|tracert|ping)\b")

# 把问题转出去的兜底措辞。BANNED_CONTENT_PATTERNS 拦的是"联系技术支持"那批词，
# 这里补的是同一件事的另一种说法——排查不出来时写"转人工分析""收集诊断信息升级处理"，
# agent 同样执行不了，排障链条到这里就断了。
# 允许的写法是：输出"未找到根因" + 已执行的检查项 + 关键回显，交用户判断。
ESCALATION_PATTERNS = [
    (r"转人工|升级处理|收集诊断信息|派单|上报工单|转专家|工单",
     "排查不出根因时不要写成把问题转出去的动作，agent 执行不了；"
     "应输出“未找到根因”、已执行的检查项与关键回显，交用户判断"),
]

# 这些关键字后面跟的是随设备而变的对象名/编号，必须写成 <参数>。
# qwen 会把样例回显里的具体值抄进修复CLI（实测抄出过 `bgp 100`、
# `undo segment-list list1`）——在别的设备上执行要么失败，要么误建/误删对象。
DEVICE_OPERAND = re.compile(
    r"^(?:undo\s+)?(bgp|segment-list|interface|peer|router-id|ospf|isis)\s+(\S+)")
POLICY_OPERAND = re.compile(r"^(?:undo\s+)?(srv6-te\s+policy|sr-te\s+policy)\s+(\S+)")
# 带数字的操作数才当成"具体实例"：`bgp route-learning` 这种子关键字不算，
# 而 policy1 / list1 / 100 / GigabitEthernet0/1/0 都带数字
LOOKS_LIKE_INSTANCE = re.compile(r"^[\w./:-]*\d[\w./:-]*$")


def inline_commands(markdown: str) -> list:
    """正文里被反引号包起来、看着像设备命令的片段。"""
    return [span for span in INLINE_CODE.findall(markdown)
            if re.match(r"^(display|ping|tracert|undo|system-view|commit|"
                        r"segment-routing|bgp|interface)\b", span.strip())]


def cli_lines(markdown: str) -> list:
    """所有像命令的行：修复方案的CLI在围栏代码块里，查询命令在行内代码里。"""
    lines = []
    for block in FENCE_BLOCK.findall(markdown):
        lines += [ln.strip() for ln in block.splitlines() if ln.strip()]
    lines += [span.strip() for span in INLINE_CODE.findall(markdown)]
    return lines


def _normalize_command(text: str) -> str:
    """去掉参数与多余空白，用于比对两条命令是不是同一条。

    还要去掉反斜杠：markdown 表格里的管道符必须转义成 `\\|`，不还原的话
    `display paf \\| include XXX` 会和表里的原命令对不上，被误判成编造的命令。
    """
    text = text.replace("\\", "")
    return re.sub(r"\s+", " ", re.sub(r"<[^>\n]*>", "", text)).strip().lower()


def known_commands(scenario: dict) -> set:
    """步骤表里出现过的查询命令（命令行列 + 修复建议列 + 修复验证列）。"""
    known = set()
    for step in scenario["steps"]:
        for field in ("command", "fix", "verify"):
            for match in re.finditer(r"(?:display|tracert|ping)[^\n，。；]*",
                                     step.get(field) or ""):
                normalized = _normalize_command(match.group(0))
                if normalized:
                    known.add(normalized)
    return known


def check_hardcoded_operands(content: str) -> str:
    """修复CLI里不许出现从样例回显抄来的具体对象名/编号。"""
    for raw in cli_lines(content):
        if QUERY_PREFIX.match(raw):
            continue   # 查询命令的过滤串（如 | include SPEC_RES_xxx）是固定字面量
        match = POLICY_OPERAND.match(raw) or DEVICE_OPERAND.match(raw)
        if not match:
            continue
        keyword, operand = match.groups()
        if operand.startswith("<"):
            continue
        if not LOOKS_LIKE_INSTANCE.match(operand):
            continue   # 跟的是子关键字（如 bgp route-learning），不是实例名
        return (f"修复CLI里 `{raw}` 的 {keyword} 操作数写成了具体值 {operand!r}——"
                f"这类取值随设备而变（多半是从样例回显里抄的），必须写成 <参数>")
    return ""


def check_unknown_commands(content: str, scenario: dict) -> str:
    """不许编造步骤表里没有的查询命令。"""
    known = known_commands(scenario)
    for raw in cli_lines(content):
        if not QUERY_PREFIX.match(raw):
            continue
        normalized = _normalize_command(raw)
        if not normalized or any(normalized == k or normalized in k or k in normalized
                                 for k in known):
            continue
        return (f"正文里的 `{raw}` 在步骤表中不存在——表没给的命令不要自己编，"
                f"没有可用命令时写到根因判定为止")
    return ""


def check_skill_format(content: str, scenario: dict) -> str:
    """校验改写结果，返回错误说明（空串表示通过）。"""
    content = (content or "").strip()
    if not content:
        return "没有生成内容"
    if not re.match(r"^---\s*\n", content):
        return "缺少frontmatter"
    # 不要求一级标题：新格式里 frontmatter 之后直接就是 `## 入参列表`，
    # 篇名已经由 frontmatter 的 name/description 承担了
    if len(re.findall(r"^\s*```", content, re.MULTILINE)) % 2 != 0:
        return "代码块围栏未闭合，疑似输出被截断"

    for pattern, advice in BANNED_CONTENT_PATTERNS + ESCALATION_PATTERNS:
        hit = re.search(pattern, content)
        if hit:
            return f"正文出现了'{hit.group(0)}'：{advice}"

    hit = TABLE_REFERENCE.search(content)
    if hit:
        return (f"正文残留了表内编号说法'{hit.group(0)}'，"
                f"agent看不到这张表，必须换成真正的命令")

    # 命令必须在反引号里。比对用"第一个参数之前的字面量部分"而不是前两个词：
    # 表里 display current-configuration configuration segment-routing-ipv6 与
    # ...configuration bgp 前两个词相同，display paf | include SPEC_RES_A 与
    # SPEC_RES_B 也是，只比前两个词的话模型漏写一条照样能通过。参数名会被改写，
    # 所以只比参数之前的部分。
    for step in scenario["steps"]:
        if not step["command"]:
            continue
        literal = step["command"].split("<")[0].strip()
        if not literal:
            continue
        if not any(literal in span for span in inline_commands(content)):
            return (f"命令 {step['command']!r} 没有以行内代码（反引号包裹）的形式"
                    f"出现在正文里")

    # 参数一律 <>，不能留 {} / [] / XXX
    for span in inline_commands(content):
        hit = BAD_PLACEHOLDER.search(span)
        if hit:
            return (f"命令 `{span}` 里的参数 {hit.group(0)!r} 不是尖括号形式，"
                    f"参数一律写成 <小写-连字符>")

    # 同一参数名在全篇必须一致：<endpoint-ipv6> 和 <endpointipv6> 不能并存
    params = {p for span in inline_commands(content)
              for p in re.findall(r"<([^>\n]+)>", span)}
    normalized = {}
    for param in params:
        key = re.sub(r"[-_\s]", "", param).lower()
        normalized.setdefault(key, []).append(param)
    for key, variants in normalized.items():
        if len(variants) > 1:
            return (f"同一个参数在正文里有多种写法: {'、'.join(sorted(variants))}，"
                    f"全篇必须统一")

    error = check_hardcoded_operands(content)
    if error:
        return error
    error = check_unknown_commands(content, scenario)
    if error:
        return error

    for check in (check_structure, check_precheck_linear, check_declared_params):
        error = check(content)
        if error:
            return error

    error = check_root_causes(content, scenario)
    if error:
        return error

    # 步骤编号要从1开始连续，"跳转步骤N"必须指向真实存在的步骤
    steps_section = split_sections(content).get("排查步骤", "")
    step_numbers = [int(n) for n, _ in STEP_HEADING.findall(steps_section)]
    if not step_numbers:
        return ("排查步骤里没有找到形如 `### 步骤 1：名称` 的步骤标题，"
                "不符合输出格式要求")
    if step_numbers != list(range(1, len(step_numbers) + 1)):
        return f"步骤编号不是从1开始的连续整数: {step_numbers}"
    for match in STEP_REFERENCE.finditer(content):
        if int(match.group(1)) not in set(step_numbers):
            return (f"正文里的“{match.group(0)}”指向了不存在的步骤"
                    f"（本篇的步骤号为 {step_numbers}）")
    return ""


FRONTMATTER_KEY = re.compile(r"^(name|description)\s*:\s*\S")


def restore_frontmatter(content: str) -> tuple:
    """把模型漏掉的 frontmatter 分隔线补回去，返回 (内容, 是否补过)。

    qwen 常把 `---` 丢掉、只留下开头的 name/description 两行（评测优化流水线上
    也踩过，见 skill_eval_optimize_pipeline.restore_frontmatter）。这不是内容
    有问题，为一对分隔线判失败、重跑一整轮改写太贵，所以直接补。

    只认开头连续的 name/description 行：遇到空行或正文就停，不会把正文里恰好
    形如 `xxx: yyy` 的句子卷进 frontmatter。
    """
    content = (content or "").strip()
    if not content or content.startswith("---"):
        return content, False
    lines = content.splitlines()
    head = []
    for line in lines:
        if FRONTMATTER_KEY.match(line):
            head.append(line)
            continue
        break
    if not any(line.startswith("name") for line in head):
        return content, False
    rest = "\n".join(lines[len(head):]).lstrip("\n")
    return "---\n" + "\n".join(head) + "\n---\n\n" + rest, True


def prepare_content(reply: str) -> tuple:
    """从回复里取出skill正文，返回 (内容, 是否补过frontmatter)。"""
    return restore_frontmatter(extract_markdown_content(reply))


def make_extractor(scenario: dict):
    """内容不合规时抛异常让 call_model_with_retry 重试，并把原因带进最终报错。"""
    def _extractor(text: str) -> str:
        try:
            extract_json_block(text)
        except (ValueError, json.JSONDecodeError) as e:
            raise ValueError(f"改写说明的json解析失败: {e}")
        content, _ = prepare_content(text)
        error = check_skill_format(content, scenario)
        if error:
            raise ValueError(f"{error}；提取到的内容开头: {content[:120]!r}")
        return text
    return _extractor


def convert_scenario(args: tuple) -> dict:
    scenario, skill_path, api_url, model_name, max_tokens, timeout = args
    prompt = (PROMPT_TEMPLATE
              .replace("<scenario>", format_scenario(scenario))
              .replace("<format_spec>", build_format_spec(scenario))
              .replace("<writing_rules>", WRITING_RULES))
    print(f"[PID {os.getpid()}] 改写: {scenario['name']}"
          f"（{len(scenario['steps'])} 步）→ {skill_path}")

    result = {"scenario": scenario["name"], "skill_path": skill_path,
              "steps": len(scenario["steps"]), "rows": scenario["rows"]}
    reply = call_model_with_retry(api_url, model_name, prompt,
                                  extractor=make_extractor(scenario),
                                  max_tokens=max_tokens, timeout=timeout)
    if reply.startswith("错误："):
        result["error"] = reply
        return result
    try:
        decision = extract_json_block(reply)
    except (ValueError, json.JSONDecodeError) as e:
        result["error"] = f"错误：改写说明解析失败: {e}"
        return result
    content, repaired = prepare_content(reply)
    result["content"] = content
    result["invalid"] = check_skill_format(content, scenario)
    result["branches_expanded"] = decision.get("branches_expanded", "")
    result["commands_normalized"] = decision.get("commands_normalized", [])
    if repaired:
        result["repaired"] = "模型没输出frontmatter的---分隔线，已补回"
    return result


def build_report(results: list, audit: list, xlsx_path: str, output_dir: str,
                 dry_run: bool) -> str:
    lines = [
        "# 排障步骤表改写为skill的说明",
        "",
        f"- 步骤表：`{xlsx_path}`",
        f"- 输出目录：`{output_dir}`",
        f"- 生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
    ]
    if dry_run:
        lines.append("- 运行模式：DRY-RUN（未写入skill文件）")

    if audit:
        lines += ["", "## 步骤表体检（需要改表）", ""]
        for name, issues in audit:
            for issue in issues:
                lines.append(f"- `{name}`：{issue}")

    lines += ["", "## 变更总览", "",
              "| 场景 | skill | 步骤数 | 结果 |", "| --- | --- | --- | --- |"]
    for r in results:
        verdict = ("失败：" + (r.get("error") or r["invalid"])
                   if r.get("error") or r.get("invalid") else "已改写")
        lines.append(f"| {r['scenario']} | `{r['skill_path']}` | "
                     f"{r['steps']} | {verdict} |")

    for r in results:
        if r.get("error") or r.get("invalid") or not r.get("content"):
            continue
        lines += ["", f"## 新建skill：`{r['skill_path']}`", "",
                  f"- 来源场景：{r['scenario']}（步骤表第{r['rows'][0]}-{r['rows'][1]}行）",
                  f"- 分支拆解：{r.get('branches_expanded', '')}"]
        if r.get("repaired"):
            lines.append(f"- 自动修复：{r['repaired']}")
        if r.get("commands_normalized"):
            pairs = "；".join(
                f"{c.get('from', '')} → {c.get('to', '')}"
                for c in r["commands_normalized"] if isinstance(c, dict))
            lines.append(f"- 参数统一：{pairs}")
        lines += ["", "<details><summary>改动内容</summary>", "", "````markdown",
                  r["content"].strip(), "````", "", "</details>"]

    failed = [r for r in results if r.get("error") or r.get("invalid")]
    if failed:
        lines += ["", "## 处理失败（未落盘，需重跑）", ""]
        for r in failed:
            lines += [f"### `{r['skill_path']}`", "",
                      f"- 失败原因：{r.get('error') or r['invalid']}"]
            if r.get("content"):
                lines += ["", "<details><summary>模型生成的内容（供排查）</summary>", "",
                          "````markdown", r["content"].strip()[:4000], "````", "",
                          "</details>", ""]
    return "\n".join(lines) + "\n"


def main(XLSX_PATH, OUTPUT_DIR, API_URL, MODEL_NAME, WORKERS, REPORT_PATH,
         SHEET_NAME=None, CATEGORY=DEFAULT_CATEGORY, DRY_RUN=True,
         CHECK_ONLY=False, MAX_TOKENS=None, TIMEOUT=600):
    print("=" * 60)
    print("排障步骤表改写为skill流水线")
    print("=" * 60)
    print(f"\n步骤表: {XLSX_PATH}")
    print(f"输出目录: {OUTPUT_DIR}")
    print(f"模型: {MODEL_NAME}")

    if not os.path.isfile(XLSX_PATH):
        print(f"错误: 步骤表不存在: {XLSX_PATH}")
        sys.exit(1)

    scenarios = parse_sheet(XLSX_PATH, SHEET_NAME)
    print(f"\n解析出 {len(scenarios)} 个故障场景:")
    audit = []
    for scenario in scenarios:
        skill_path = derive_skill_path(scenario, CATEGORY)
        print(f"  ● {scenario['name']}（sheet {scenario['sheet']} "
              f"第{scenario['rows'][0]}-{scenario['rows'][1]}行，"
              f"{len(scenario['steps'])} 步）→ {skill_path}")
        issues = audit_commands(scenario) + audit_step_numbers(scenario)
        if issues:
            audit.append((scenario["name"], issues))
            for issue in issues:
                print(f"      [WARN] {issue}")

    if audit:
        print(f"\n[WARN] {sum(len(i) for _, i in audit)} 处步骤表问题，"
              f"建议先改表再改写——ragIndex 重号会让改写出的skill指向错的命令")

    # 改写要花模型调用，先用 --check 确认表解析得对、体检没问题
    if CHECK_ONLY:
        print(f"\n--check：{len(scenarios)} 个场景解析完成，"
              f"{sum(len(i) for _, i in audit)} 处表问题（未调用模型，未写任何文件）")
        sys.exit(1 if audit else 0)

    task_args = [(s, derive_skill_path(s, CATEGORY), API_URL, MODEL_NAME,
                  MAX_TOKENS, TIMEOUT) for s in scenarios]
    with Pool(processes=WORKERS) as pool:
        results = pool.map(convert_scenario, task_args)

    print("\n" + "=" * 60)
    for r in results:
        if r.get("error") or r.get("invalid"):
            print(f"[FAIL] {r['skill_path']}: {r.get('error') or r['invalid']}")
            continue
        path = os.path.join(OUTPUT_DIR, *r["skill_path"].split("/"))
        if DRY_RUN:
            print(f"[DRY-RUN] create → {path}")
            continue
        os.makedirs(os.path.dirname(path), exist_ok=True)
        with open(path, 'w', encoding='utf-8') as f:
            f.write(r["content"].strip() + "\n")
        print(f"已新建: {path}")

    report = build_report(results, audit, XLSX_PATH, OUTPUT_DIR, DRY_RUN)
    os.makedirs(os.path.dirname(os.path.abspath(REPORT_PATH)), exist_ok=True)
    with open(REPORT_PATH, 'w', encoding='utf-8') as f:
        f.write(report)
    print(f"\n改写说明已保存到: {REPORT_PATH}")
    if DRY_RUN:
        print("这是DRY-RUN；审过报告后用 "
              f"python3 apply_change_report.py {REPORT_PATH} {OUTPUT_DIR} --apply 落盘。")

    failed = [r for r in results if r.get("error") or r.get("invalid")]
    if failed:
        print(f"\n存在 {len(failed)} 个处理失败项，请查看改写说明后重跑")
        sys.exit(1)


def validate_files(paths: list, xlsx_path: str, sheet_name: str = None) -> int:
    """对已有的skill文件单独跑一遍格式校验，不调模型。

    校验器平时只在流水线内部对模型回复生效；落盘之后想再验一遍（比如手工改过
    正文、或想确认某篇老skill合不合新规范）就需要这个入口。
    命令是否齐全要对着步骤表比，所以仍要读表。
    """
    scenarios = parse_sheet(xlsx_path, sheet_name)
    by_path = {derive_skill_path(s): s for s in scenarios}
    failures = 0
    for path in paths:
        content = open(path, 'r', encoding='utf-8', errors='replace').read()
        # 按文件名找对应场景；只有一个场景时直接用它
        key = next((k for k in by_path if os.path.basename(k) == os.path.basename(path)),
                   None)
        scenario = by_path[key] if key else (scenarios[0] if len(scenarios) == 1 else None)
        if scenario is None:
            print(f"[SKIP] {path}: 在步骤表里找不到同名场景，无法核对命令是否齐全")
            continue
        error = check_skill_format(content, scenario)
        if error:
            failures += 1
            print(f"[FAIL] {path}\n       {error}")
        else:
            print(f"[OK]   {path}（对照场景「{scenario['name']}」）")
    print(f"\n共 {len(paths)} 个文件，失败 {failures} 个")
    return 1 if failures else 0


if __name__ == "__main__":
    # --validate <文件…>：单独校验已有的skill文件，不调模型
    if "--validate" in sys.argv[1:]:
        targets = [a for a in sys.argv[1:] if a != "--validate" and not a.startswith("-")]
        if not targets:
            print("用法: python3 excel_skill_distill_pipeline.py --validate <skill.md> [更多文件…]")
            sys.exit(2)
        sys.exit(validate_files(targets, "excel_cases/排障步骤表.xlsx"))

    main(
        XLSX_PATH="excel_cases/排障步骤表.xlsx",
        OUTPUT_DIR=f"skills_from_excel/{datetime.now().strftime('%m-%d')}",
        # 地址与密钥从 .env 按模型名解析（见 model_config.py）
        API_URL=None,
        # 改写要拆分支、统一参数、还要顾及跨步骤的命令复用。这个任务是"照着已有判据
        # 重写"而不是"从散文里提炼判据"，未必吃推理，所以用调用量大、重跑便宜的
        # qwen3.6-27b；跑过几轮后可用 compare_models.py 和 MiniMax-M2.7 对比复核。
        # 注意换模型也换了输出预算：qwen 那档是 16384，MiniMax 是 32768，
        # 场景步骤多时若撞上 finish_reason=length，用下面的 MAX_TOKENS 单独调大。
        MODEL_NAME="qwen3.6-27b",
        WORKERS=3,
        REPORT_PATH=f"reports/excel_skill_report_{datetime.now().strftime('%m-%d')}.md",
        SHEET_NAME=None,
        CATEGORY=DEFAULT_CATEGORY,
        DRY_RUN=True,
        # 加 --check 只解析和体检，不调模型
        CHECK_ONLY="--check" in sys.argv[1:],
        MAX_TOKENS=None,
        TIMEOUT=600,
    )
