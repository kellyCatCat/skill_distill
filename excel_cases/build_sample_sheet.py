#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""生成排障步骤表的样例 xlsx。

xlsx 是二进制，进了 git 就看不出 diff，所以表的内容以这个脚本为准：改内容改这里
再重跑，评审时看这个文件而不是去开 Excel。

表的结构（`skill_excel_distill_pipeline` 按这个结构解析）：
  - 第 1 行为表头，11 列，列名连同列里那几句填写说明一起保留（原表就是这么写的）；
  - 一个 sheet 放多个故障场景，靠 A、B 两列的**合并单元格**分块：一个合并区间
    = 一个场景，区间覆盖的行就是该场景的排障步骤；
  - 合并区间的值只写在左上角那一格（openpyxl 的其余格是只读的 MergedCell）。

用法：
  python3 excel_cases/build_sample_sheet.py [输出路径]
"""
import os
import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill
from openpyxl.utils import get_column_letter

DEFAULT_OUTPUT = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                              "排障步骤表.xlsx")

SHEET_NAME = "排障步骤"

# 表头照抄原表：列名底下那几句是给填表人的说明，也是原样保留的一部分
HEADERS = [
    "排障目标\n1、有告警或者事件写进来\n2、提供一种故障构造方法",
    "组网场景\n按需，协议类可能涉及对端，截图放进来即可",
    "排障步骤编号",
    "排障步骤描述\n",
    "步骤详细描述",
    "本步骤需要使用的命令行的编号及使用目的（ragIndex）",
    "命令行",
    "回显",
    "配置修复建议，定位到根因的步骤里给出，如果需要人工修复就给出建议； "
    "如果可以自修复（配置错误）给出CLI，包括CLI生成规则和参数分配规则，细化到具体参数。",
    "修复建议影响性\n修改配置的，如果影响性大，则需要可靠性保障<比如仿真>",
    "修复验证，怎么验证",
]

# 列宽按内容类型给：判据类的列要宽，回显更宽
COLUMN_WIDTHS = [30, 18, 10, 24, 60, 28, 44, 76, 34, 20, 34]

# 等宽字体的列（0基）：命令行与回显是终端文本，列对齐靠空格，
# 用比例字体会让回显里的字段列全部错位，判据又恰好依赖这些字段
MONOSPACE_COLUMNS = {6, 7}

TEXT_FONT = "Arial"
MONO_FONT = "Consolas"


SRV6_GOAL = """SRv6 TE Policy down告警
构造方法：1、静态配置policy场景candidate path/segment list配置不全
2、BGP动态下发场景，未配置policy地址族
3、srlist的sid不可达，如链路故障"""

ECHO_STEP1 = """[~HUAWEI]display srv6-te policy endpoint 1::1 color 100
PolicyName : policy1
Color                   : 100                            Endpoint             : 1::1
TunnelId                : 16385
TunnelType              : SRv6-TE Policy                 DelayTimerRemain     : -
Policy State            : Down (Init)                    State Change Time    : 2023-08-01 13:05:32
Admin State             : Up                             Traffic Statistics   : Disable
Backup Hot-Standby      : Disable                        BFD                  : Disable
Interface Index         : -                              Interface Name       : -
Interface State         : -                              Encapsulation Mode   : -
Binding SID             : -
Candidate-path Count    : 1

 Candidate-path Preference : 100
 Path State             : Inactive (Invalid)             Path Type            : -
 Protocol-Origin        : Configuration(30)              Originator           : 0, 0.0.0.0
 Discriminator          : 100                            Binding SID          : -
 GroupId                : 32769                          Policy Name          : policy1
 Template ID            : 0                              Path Verification    : Enable
 DelayTimerRemain       : -                              Network Slice ID     : -
 Segment-List Count     : 1
  Segment-List          : list1
   Segment-List ID      : 114689                         XcIndex              : -
   List State           : Down (SID Stack Down)          DelayTimerRemain     : -
   Verification State   : SID Unreachable                SuppressTimeRemain   : -
   PMTU                 : 9600                           Active PMTU          : 9600
   Weight               : 1                              BFD State            : -
   Loop Detection State : Up                             BFD Delay Remain     : -
   Network Slice ID     : -
   Binding SID          : -
   Reverse Binding SID  : -
   SID :
         A1:1::1
         B1:1::1"""

ECHO_STEP2 = """[~HUAWEI]display current-configuration configuration segment-routing-ipv6
#
segment-routing ipv6
 segment-list 1
  index 1 sid ipv6 1::1
 srv6-te policy a endpoint 1::1 color 1
  candidate-path preference 100
   segment-list 1"""

ECHO_STEP3 = """[~HUAWEI]display current-configuration configuration bgp
#
bgp 100
 router-id 1.1.1.1
 private-4-byte-as enable
 peer 2.2.2.2 as-number 100
 peer 2.2.2.2 connect-interface LoopBack0
 peer 1::2 as-number 100
 peer 1::2 connect-interface LoopBack0
 #
 ipv4-family unicast
  undo synchronization
  peer 2.2.2.2 enable
 #
 ipv4-family sr-policy
  router-id filter
  peer 2.2.2.2 enable
  peer 2.2.2.2 route-policy rt export
  peer 2.2.2.2 advertise-ext-community
 #
 ipv6-family sr-policy
  router-id filter
  peer 2.2.2.2 enable
  peer 2.2.2.2 route-policy rt export
  peer 2.2.2.2 advertise-ext-community
  peer 1::2 enable
#
bgp route-learning acceleration enable
#
return"""

ECHO_STEP5 = """[HUAWEI] display bfd session srv6-segment-list 5
(w): State in WTR
(*): State is invalid
--------------------------------------------------------------------------------
Local      Remote     PeerIpAddr      State     Type        InterfaceName
--------------------------------------------------------------------------------
16395      100000000  xxx::xxx
                                      Down        D_SID_LIST        -
--------------------------------------------------------------------------------"""

ECHO_STEP6 = """[HUAWEI] display srv6-te policy source-sid
SID          : 2::1:0:7
FuncType     : end-x        Topology     : 2
Overload     : true         S-Flag       : 0
  Producer   : ISIS, Process 1, Level-1
    Node     : 0000.0000.0002.0000
    Local    : 0000.0000.0002.0000
    Peer     : 0000.0000.0002.0100
  Producer   : ISIS, Process 1, Level-2
    Node     : 0000.0000.0002.0000
    Local    : 0000.0000.0002.0000
    Peer     : 0000.0000.0002.0100"""

ECHO_STEP7 = """[~HUAWEI] display paf | include SPEC_RES_SRV6POLICY_MAX_NUM
Info: It will take a long time if the content you search is too much or the string you input is too long, you can press CTRL_C to break.
----------------------------------------------------------------------------------------------------------------------
PafName                                                             PafValue             Description
----------------------------------------------------------------------------------------------------------------------
SPEC_RES_SRV6POLICY_MAX_NUM                                         32768                The max number of srv6-te policies
----------------------------------------------------------------------------------------------------------------------"""

ECHO_STEP8 = """[~HUAWEI] display paf | include SPEC_RES_SRV6POLICY_SEGLIST_GLOBAL_NUM
Info: It will take a long time if the content you search is too much or the string you input is too long, you can press CTRL_C to break.
----------------------------------------------------------------------------------------------------------------------
PafName                                                             PafValue             Description
----------------------------------------------------------------------------------------------------------------------
SPEC_RES_SRV6POLICY_SEGLIST_GLOBAL_NUM                              65536                The max number of srv6-segment-lists globally
----------------------------------------------------------------------------------------------------------------------"""

# 每个场景一块：goal/topology 写进合并单元格，steps 是该场景的排障步骤行。
# 每行 9 个字段，对应表头第 3~11 列。
SCENARIOS = [
    {
        "goal": SRV6_GOAL,
        "topology": "",
        "steps": [
            [
                1,
                "检查隧道状态",
                "通过执行1号命令行查询SRv6 TE Policy，判断指定endpoint/color的SRv6 TE Policy"
                "是否存在，如果不存在则返回“SRv6 TE Policy {endpoint/color}不存在”，并结束执行；"
                "如果存在，则判断此SRv6 TE Policy的Policy State字段是否为Up，如果不是Up，"
                "则继续向下执行；如果Policy State字段是Up，则判断srlist的状态List State是否为Up，"
                "如果不是，则继续向下执行；如果都是Up，则返回“SRv6 TE Policy {endpoint/color}"
                "状态正常”并结束执行。",
                "1：查询SRv6 TE Policy",
                "display srv6-te policy endpoint <endpointipv6> color <colorid>",
                ECHO_STEP1,
                "",
                "",
                "",
            ],
            [
                2,
                "静态配置的SRv6 TE Policy是否配置完整",
                "通过执行2号命令行，查看segment-routing ipv6下是否配置SRv6 TE Policy、"
                "candidate path，如果没有配置进行第三步检查；如果有candidate path配置，"
                "检查是否引用和segment list并且segment list下有SID配置，如果没有配置，"
                "则返回“SRv6 TE Policy配置不完整”",
                "2: 查询静态SRv6 TE Policy配置",
                "display current-configuration configuration segment-routing-ipv6",
                ECHO_STEP2,
                "按照到SRv6 TE Policy的endpoint的路径，在segment-list下配置SID栈",
                "",
                "",
            ],
            [
                3,
                "BGP动态下发场景的SRv6 TE Policy是否配置完整",
                "通过执行3号命令行，查看bgp下是否配置ipv6-family sr-policy地址族，"
                "如果没有配置，则提示“ipv6-family sr-policy地址族未配置”并结束执行，"
                "否则继续向下执行",
                "3: 查询静态sr-policy地址族配置",
                "display current-configuration configuration bgp",
                ECHO_STEP3,
                "BGP视图下配置ipv6-family sr-policy",
                "",
                "查看sr-policy peer是否建立成功，状态为Established，"
                "命令行display bgp sr-policy ipv6 peer ",
            ],
            [
                4,
                "检查SRv6 TE Policy是否被shutdown",
                "检查1号命令行的执行回显中Policy State是否为Down (Shutdown)，如果是，"
                "则返回“SRv6 TE Policy被shutdown”，否则继续向下执行",
                "",
                "",
                "",
                "进入 endpoint xxx color xxx下\nsrpolicy视图下执行undo shutdown",
                "",
                "",
            ],
            [
                5,
                "检查是否bfd down导致中断",
                "检查1号命令行的执行回显中srlist的BFD State是否为Down，如果不是，跳过本步，"
                "否则执行4号命令查看具体的bfd状态，提示错误信息“bfd检测Down”并结束运行，"
                "否则继续执行。",
                "4: 检查srlist的bfd状态",
                "display bfd session srv6-segment-list <segment-list-id>",
                ECHO_STEP5,
                "执行tracert srv6-te policy endpoint-ip <endpoint-ip> color <color>，"
                "查看哪一跳不可达，排查链路状态",
                "",
                "",
            ],
            [
                6,
                "检查是否故障感知down导致中断",
                "检查1号命令行的执行回显中srlist的List State是否为Down (SID Stack Down)"
                "并且Verification State是否为SID Unreachable，如果不是，跳过本步，"
                "否则执行6号命令查看ISIS拓扑是否存在srlist中的sid，提示错误信息"
                "“故障感知检测Down”并结束运行，否则继续执行。",
                "6、检查拓扑数据",
                "display srv6-te policy source-sid",
                ECHO_STEP6,
                "",
                "",
                "",
            ],
            [
                7,
                "SRv6 TE Policy超限",
                "检查1号命令行的执行回显中Policy State是否为Down (Overrun)，如果不是，"
                "跳过本步，否则执行5号命令查看设备支持的规格，提示错误信息"
                "“SRv6 TE Policy超限”并结束运行，否则继续执行。",
                "5: 检查支持的SRv6 TE Policy规格",
                "display paf | include SPEC_RES_SRV6POLICY_MAX_NUM",
                ECHO_STEP7,
                "减少policy数量",
                "",
                "",
            ],
            [
                8,
                "SRv6 TE Policy srlist超限",
                "检查1号命令行的执行回显中List State是否为Down (Overrun)，如果不是，"
                "跳过本步，否则执行6号命令查看设备支持的规格，提示错误信息“srlist超限”"
                "并结束运行。",
                "6: 检查支持的SRv6 TE Policy规格",
                "display paf | include SPEC_RES_SRV6POLICY_SEGLIST_GLOBAL_NUM",
                ECHO_STEP8,
                "减少srlist数量",
                "",
                "",
            ],
        ],
    },
]


def estimate_row_height(values) -> float:
    """按单元格里最多的行数估行高，并封顶。

    回显动辄三十多行，让 Excel 自动撑开会把一行拉到整屏、表根本没法翻。封顶之后
    超出的部分双击行边界即可展开，不影响内容本身。
    """
    lines = 1
    for value in values:
        if isinstance(value, str) and value:
            lines = max(lines, value.count("\n") + 1)
    return min(15.0 * lines, 300.0)


def build_workbook() -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME

    header_font = Font(name=TEXT_FONT, bold=True, size=10)
    header_fill = PatternFill("solid", fgColor="DDEBF7")
    header_align = Alignment(wrap_text=True, vertical="center", horizontal="center")

    for col, title in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        ws.column_dimensions[get_column_letter(col)].width = COLUMN_WIDTHS[col - 1]
    ws.row_dimensions[1].height = 60

    row = 2
    for scenario in SCENARIOS:
        start_row = row
        for step in scenario["steps"]:
            for offset, value in enumerate(step):
                col = 3 + offset
                cell = ws.cell(row=row, column=col, value=value if value != "" else None)
                cell.font = Font(
                    name=MONO_FONT if (col - 1) in MONOSPACE_COLUMNS else TEXT_FONT,
                    size=9 if (col - 1) in MONOSPACE_COLUMNS else 10)
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            ws.row_dimensions[row].height = estimate_row_height(step)
            row += 1
        end_row = row - 1

        # 合并单元格只写左上角：其余格是只读的 MergedCell
        for col, value in ((1, scenario["goal"]), (2, scenario["topology"])):
            anchor = ws.cell(row=start_row, column=col, value=value or None)
            anchor.font = Font(name=TEXT_FONT, size=10)
            anchor.alignment = Alignment(wrap_text=True, vertical="top")
            if end_row > start_row:
                ws.merge_cells(start_row=start_row, start_column=col,
                               end_row=end_row, end_column=col)

    # 冻结表头与场景两列，翻到第 8 步时还能看见这一步属于哪个场景
    ws.freeze_panes = "C2"
    return wb


def main(output_path: str):
    wb = build_workbook()
    os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)
    wb.save(output_path)

    ws = wb[SHEET_NAME]
    total_steps = sum(len(s["steps"]) for s in SCENARIOS)
    print(f"已生成: {output_path}")
    print(f"  sheet: {ws.title}｜{len(HEADERS)} 列｜"
          f"{len(SCENARIOS)} 个场景｜{total_steps} 个排障步骤")
    for rng in ws.merged_cells.ranges:
        print(f"  合并单元格: {rng}")


if __name__ == "__main__":
    main(sys.argv[1] if len(sys.argv) > 1 else DEFAULT_OUTPUT)
