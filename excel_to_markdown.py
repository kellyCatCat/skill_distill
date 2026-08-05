#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""把排障步骤表（Excel）转成蒸馏流水线能吃的markdown源文档。

排障步骤表是一行一步、列里塞多行文本的形态：命令行、回显、修复建议各占一列，
单元格内部还带"1、…2、…"的子编号。直接把表格转成markdown表格没用——回显动辄
几十行，塞进表格单元格既读不了、模型也抓不住结构。所以这里按**一行一小节**展开，
命令行与回显放进围栏代码块，其余列按固定顺序渲染成加粗小标题。

每个sheet转成一篇markdown，文件名即sheet名。转出来的是**源文档**不是skill：
它没有frontmatter，要交给 skill_self_distill_pipeline.py 蒸馏成skill。

表里有"排障目标"这类场景列时，一张表可以含多个故障场景（该格非空就是新场景的
开始，后续步骤行留空表示接着上一个场景）：此时按场景分节、步骤降为三级标题，
否则第二个场景的步骤会和第一个场景的平铺在同一层，蒸馏时两条排查链会混成一条。

输入支持两种：
  - .xlsx / .xlsm（需要 openpyxl：pip install openpyxl）
  - .csv / .tsv（从Excel直接复制粘贴到文本文件就是这个形态，不装依赖也能跑）

用法：
  python3 excel_to_markdown.py 步骤表.xlsx
  python3 excel_to_markdown.py 步骤表.xlsx docs_from_excel/BFD
  python3 excel_to_markdown.py 步骤表.xlsx --sheet BFD会话异常 --title "BFD会话异常排障步骤"
  # 直接摆成蒸馏源树的布局（<一级目录>/<二级目录>/<sheet>.md），转完就能跑蒸馏
  python3 excel_to_markdown.py 步骤表.xlsx result/新来源 --tree "故障处理：网络可靠性/BFD故障案例"
  python3 excel_to_markdown.py 步骤表.xlsx --skip-unsupported   # 丢掉标"不支持"的行
  python3 excel_to_markdown.py 步骤表.xlsx --stdout             # 只打印不写文件
  python3 excel_to_markdown.py 步骤表.xlsx --style branch       # 分支格式

--style 有两种排版，内容完全相同、只是层级不同：
  section（默认）一步一个 `## 步骤N …` 小节，各列作为加粗小标题；
  branch        一步一个有序列表项、判断分支作为子项，形如

      1. **检查接口物理及协议状态**

         在任意视图下执行 `display interface ...` 检查入接口是否收到数据报文。
         - 若接口状态为 Down，请检查线路连接及接口是否被 shutdown。
         - 若接口状态为 Up 但无流量，继续后续步骤。

分支格式是**纯规则重排**：整行以"若/如果/当"开头的归为分支子项，以"1、"开头的
归为子步骤，命令加反引号。不合并、不拆句、不改一个字——这件事一旦交给模型做，
它必然顺手改写措辞，而这里的前提是保住原有的命令与描述。
"""
import csv
import os
import re
import sys

# 列名 → 字段的映射规则，**按这个顺序匹配，先命中的先占**。
# 顺序不是随便排的：列名之间互相包含，按错顺序会串列——
#   "本步骤需要使用的命令行的编号及使用目的（ragIndex）" 含"命令行"，得排在"命令行"之前；
#   "修复建议影响性…" 含"修复建议"，得排在"修复建议"之前。
FIELD_RULES = [
    # 场景级字段（不是这一步的属性，而是这一段步骤共同归属的故障场景）：
    # 排障目标那列里带着告警名和故障构造方法，一个sheet里可以有多个场景，
    # 后续步骤行的这一格是空的（合并单元格），空值表示"接着上一个场景"。
    ("goal",        ("排障目标", "故障场景", "排障场景", "故障定义")),
    ("topology",    ("组网场景", "组网")),
    ("step_no",     ("步骤编号",)),
    ("cmd_purpose", ("ragindex", "命令行的编号")),
    ("impact",      ("影响性",)),
    ("version",     ("是否支持", "eos")),
    ("verify",      ("修复验证", "怎么验证")),
    ("fix",         ("修复建议", "修复方案")),
    ("detail",      ("详细描述",)),
    ("step_title",  ("步骤描述", "步骤名称")),
    ("cmd",         ("命令行", "命令")),
    ("output",      ("回显",)),
]

# 渲染顺序与小标题（与匹配顺序无关，这里是给人和模型读的先后）
RENDER_ORDER = [
    ("detail",      "操作说明"),
    ("cmd_purpose", "命令用途"),
    ("cmd",         "命令行"),
    ("output",      "回显示例"),
    ("fix",         "修复建议"),
    ("impact",      "影响性"),
    ("verify",      "修复验证"),
    ("version",     "版本支持"),
]

# 这两列是设备输入输出，必须进围栏代码块：回显里有 # 开头的行（配置视图提示符），
# 不包起来会被markdown当成标题，整篇结构就乱了。
CODE_FIELDS = ("cmd", "output")

# 场景级字段：属于整段步骤而不是某一步，不参与步骤小节的渲染
SCENARIO_FIELDS = ("goal", "topology")

# 表头可能不在第一行（上面常压着标题行、说明行），在前若干行里找命中列名最多的那行
HEADER_SEARCH_ROWS = 10

LINK_PATTERN = re.compile(r"https?://\S+")
FILENAME_UNSAFE = re.compile(r'[\\/:*?"<>|\r\n\t]')

# --style branch 用：单元格里以"若/如果/当…"开头的一行是一个判断分支，
# 以"1、""2."开头的一行是这一步内部的子步骤。只按**整行**判断，不去切句子里的
# 分句——把一句话拆开就等于在改写描述了，而这里的前提是一个字都不改。
BRANCH_LINE = re.compile(r"^(?:若|如果|当|假如|倘若|否则)")
NUMBERED_LINE = re.compile(r"^(\d+)\s*[、.)．]\s*")

# 命令加反引号用：命令是ASCII、周围正文是中文，所以一条 display 命令一定在第一个
# 中文字符处结束。只认 display 开头的查询命令，配置命令由"命令行"列给出的原文匹配，
# 不去猜正文里哪一段是命令。
DISPLAY_IN_TEXT = re.compile(r"display[ \t][\x20-\x7E]*")


def normalize_header(text: str) -> str:
    """列名规整：去掉换行、空格、引号，转小写。

    表头里换行是常态（"修复建议影响性\\n修改配置的，如果影响性大…"），不规整就匹配不上。
    """
    return re.sub(r"[\s\"'）()（]", "", str(text or "")).lower()


def cell_text(value) -> str:
    """单元格 → 纯文本。数字要去掉Excel带来的.0，否则步骤编号会变成"1.0"。"""
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    text = str(value).replace("\r\n", "\n").replace("\r", "\n").strip()
    # 行尾空格会让markdown把它当成换行标记，统一清掉
    text = "\n".join(line.rstrip() for line in text.split("\n"))
    # 单元格里常有连着的空行，压成一个
    return re.sub(r"\n{3,}", "\n\n", text).strip()


def map_columns(header_cells: list) -> tuple:
    """表头 → ({字段: 列号}, [(列号, 原列名)])。

    没能映射到已知字段的列不丢弃，作为"其它信息"原样输出——步骤表各版本列并不统一，
    静默丢列会让源文档缺内容，而蒸馏出来之后已经看不出少了什么。
    """
    mapping, extras = {}, []
    for idx, raw in enumerate(header_cells):
        name = normalize_header(raw)
        if not name:
            continue
        # 同一字段只认第一列（已占用的字段跳过），重复列一并进"其它信息"
        hit = next((field for field, keywords in FIELD_RULES
                    if field not in mapping and any(kw in name for kw in keywords)), None)
        if hit:
            mapping[hit] = idx
        else:
            extras.append((idx, " ".join(str(raw).split())))
    return mapping, extras


def find_header_row(rows: list) -> int:
    """找出表头行号：前若干行里能映射出最多已知字段的那行。"""
    best_idx, best_hits = 0, 0
    for idx, row in enumerate(rows[:HEADER_SEARCH_ROWS]):
        hits = len(map_columns(row)[0])
        if hits > best_hits:
            best_idx, best_hits = idx, hits
    if best_hits < 2:
        raise ValueError(
            "前 %d 行里找不到表头（识别出的列少于2个）。请确认表头行包含"
            "'排障步骤编号'/'命令行'/'回显'这类列名" % HEADER_SEARCH_ROWS)
    return best_idx


def fence_for(text: str) -> str:
    """按内容里最长的反引号串决定围栏长度，避免回显里的``` 提前闭合围栏。"""
    longest = max((len(m) for m in re.findall(r"`+", text)), default=0)
    return "`" * max(3, longest + 1)


def parse_rows(rows: list) -> tuple:
    """表格行 → ([场景字典], [(列号, 列名)] 其它列, 表头行号)。

    场景字典为 {title, goal, topology, steps}。一张表里可以有多个故障场景：
    "排障目标"这一格非空就是一个新场景的开始，后面的步骤行该格为空（合并单元格），
    表示接着上一个场景。表里没有这一列时，全部步骤归入一个无名场景，输出退回到
    "整篇就是一串步骤"的形态。
    """
    header_idx = find_header_row(rows)
    mapping, extras = map_columns(rows[header_idx])

    scenarios = []
    last_no = ""
    for row in rows[header_idx + 1:]:
        values = {field: cell_text(row[idx]) if idx < len(row) else ""
                  for field, idx in mapping.items()}
        other = [(name, cell_text(row[idx]) if idx < len(row) else "")
                 for idx, name in extras]
        if not any(values.values()) and not any(v for _, v in other):
            continue                          # 整行为空（表格底部的空行）

        goal = values.pop("goal", "")
        topology = values.pop("topology", "")
        # 排障目标非空 = 新场景开始；场景换了，步骤编号也重新起算
        if goal or not scenarios:
            scenarios.append({
                "title": goal.split("\n")[0].strip(),
                "goal": goal,
                "topology": topology,
                "steps": [],
            })
            last_no = ""
        elif topology and not scenarios[-1]["topology"]:
            scenarios[-1]["topology"] = topology

        # 编号留空一般是上一步的续行，沿用上一个编号，免得小节标题变成"步骤 查看…"
        if values.get("step_no"):
            last_no = values["step_no"]
        else:
            values["step_no"] = last_no
        values["_other"] = [(n, v) for n, v in other if v]
        scenarios[-1]["steps"].append(values)

    return [s for s in scenarios if s["steps"]], extras, header_idx


def step_labels(steps: list) -> list:
    """每一步的小节名，如 '步骤1 查看BFD会话配置（情形2）'。

    同一个步骤编号出现多次是这类表的常态（同一步在不同根因下判据不同），此时补上
    "（情形N）"，否则一篇文档里会出现多个同名小节，蒸馏时分不清是哪一支。总览与
    小节标题共用这份标签，两处才对得上。
    """
    counts = {}
    for step in steps:
        counts[step.get("step_no", "")] = counts.get(step.get("step_no", ""), 0) + 1

    labels, seen = [], {}
    for seq, step in enumerate(steps, 1):
        no = step.get("step_no") or str(seq)
        seen[no] = seen.get(no, 0) + 1
        # 步骤描述本身可能是多行的，标题只取第一行，其余并入正文
        title = (step.get("step_title") or "排障步骤").split("\n")[0].strip()
        suffix = f"（情形{seen[no]}）" if counts.get(step.get("step_no", ""), 0) > 1 else ""
        labels.append(f"步骤{no} {title}{suffix}")
    return labels


def render_steps(steps: list, labels: list, heading: str) -> list:
    """渲染一组步骤的小节。heading 为 '##' 或 '###'（场景之下要降一级）。"""
    lines = []
    for step, label in zip(steps, labels):
        lines += [f"{heading} {label}", ""]

        # 步骤描述是多行时，第一行进了标题，剩下的作为正文补在最前面
        extra_title_lines = "\n".join(
            (step.get("step_title") or "").split("\n")[1:]).strip()
        if extra_title_lines:
            lines += [extra_title_lines, ""]

        for field, field_label in RENDER_ORDER:
            value = step.get(field, "")
            if not value:
                continue
            lines += [f"**{field_label}**", ""]
            if field in CODE_FIELDS:
                fence = fence_for(value)
                lines += [fence, value, fence, ""]
            else:
                lines += [value, ""]

        for name, value in step.get("_other", ()):
            lines += [f"**{name}**", "", value, ""]
    return lines


def split_commands(cell: str) -> list:
    """"命令行"单元格 → 命令列表，去掉"1、""2."这类行内编号。"""
    commands = []
    for line in (cell or "").split("\n"):
        line = NUMBERED_LINE.sub("", line.strip()).strip()
        if line:
            commands.append(line)
    return commands


def wrap_commands(text: str, commands: list) -> str:
    """给正文里的命令加上反引号，**只加标记不改字**。

    两个来源：一是"命令行"列给出的原文（精确匹配，不猜）；二是正文里以 display
    开头的查询命令——命令是ASCII、正文是中文，所以一定在第一个中文字符处结束。
    单元格原文里本来就有反引号时整段跳过，避免套两层。
    """
    if not text or "`" in text:
        return text

    spans = []
    for command in sorted(commands, key=len, reverse=True):
        start = text.find(command)
        while start != -1:
            spans.append((start, start + len(command)))
            start = text.find(command, start + len(command))
    for match in DISPLAY_IN_TEXT.finditer(text):
        end = match.end()
        while end > match.start() and text[end - 1] in " \t,;.，":
            end -= 1                      # 命令后面紧跟的空格和标点不算命令的一部分
        spans.append((match.start(), end))

    merged = []
    for start, end in sorted(spans):
        if merged and start < merged[-1][1]:
            continue                      # 与已选片段重叠（长命令优先），丢掉短的
        merged.append((start, end))

    for start, end in reversed(merged):
        marked = text[start:end].strip()
        if marked:
            text = f"{text[:start]}`{marked}`{text[end:]}"
    return text


def bullet_block(label: str, value: str, indent: str) -> list:
    """把"标签：多行取值"渲染成一个列表项。

    续行必须再缩进两格对齐到列表项的正文列（"- "占两格），否则单元格里的
    "1、…\\n2、…"第二行会掉出列表项，变成同级的另一段。
    """
    first, *rest = value.split("\n")
    lines = [f"{indent}- {label}：{first}".rstrip()]
    lines += [f"{indent}  {line}".rstrip() for line in rest]
    return lines


def render_steps_branch(steps: list, labels: list) -> list:
    """分支格式：一步一个有序列表项，判断分支作为子项。

        1. **检查接口物理及协议状态**

           在任意视图下执行 `display interface ...` 检查入接口是否收到数据报文。
           - 若接口状态为 Down，请检查线路连接。

    只做重排与加标记：单元格里的每一行原样搬过来，不合并、不拆句、不改措辞——
    要求就是"不修改已有的命令、描述"，一旦交给模型改写就守不住这条。
    """
    lines = []
    for step, label in zip(steps, labels):
        # label 形如"步骤1 查看组件的CPU占用率（情形2）"，编号提出来当列表序号
        no_and_title = label.split(" ", 1)
        marker = f"{no_and_title[0].replace('步骤', '')}. "
        indent = " " * len(marker)
        title = no_and_title[1] if len(no_and_title) > 1 else label
        lines += [f"{marker}**{title}**", ""]

        commands = split_commands(step.get("cmd", ""))
        detail = step.get("detail", "")
        # 步骤描述是多行时，第一行已经在标题里，其余并进正文
        rest_title = "\n".join((step.get("step_title") or "").split("\n")[1:]).strip()
        if rest_title:
            detail = f"{rest_title}\n{detail}" if detail else rest_title

        used_in_body = []
        for raw in detail.split("\n"):
            line = raw.strip()
            if not line:
                continue
            used_in_body.append(line)
            marked = wrap_commands(line, commands)
            numbered = NUMBERED_LINE.match(line)
            if numbered:
                # "1、xxx" → 有序子项，编号保留，只把顿号换成markdown的点
                body = wrap_commands(NUMBERED_LINE.sub("", line), commands)
                lines.append(f"{indent}{numbered.group(1)}. {body}")
            elif BRANCH_LINE.match(line):
                lines.append(f"{indent}- {marked}")
            else:
                lines.append(f"{indent}{marked}")
        lines.append("")

        body_text = "\n".join(used_in_body)
        # 正文里没出现过的命令单独列出来，否则"命令行"列的内容就丢了
        missing = [c for c in commands if c not in body_text]
        if missing:
            lines.append(f"{indent}- 命令：" + "、".join(f"`{c}`" for c in missing))

        purpose = step.get("cmd_purpose", "")
        if purpose and purpose.strip() != title.split("（")[0].strip():
            lines += bullet_block("命令用途", purpose, indent)
        for field, field_label in (("fix", "修复建议"), ("impact", "影响性"),
                                   ("verify", "修复验证"), ("version", "版本支持")):
            value = step.get(field, "")
            if value:
                lines += bullet_block(field_label, value, indent)
        for name, value in step.get("_other", ()):
            lines += bullet_block(name, value, indent)

        output = step.get("output", "")
        if output:
            fence = fence_for(output)
            lines += ["", f"{indent}回显示例：", ""]
            lines += [f"{indent}{line}".rstrip()
                      for line in f"{fence}\n{output}\n{fence}".split("\n")]
        lines.append("")
    return lines


def render_document(title: str, scenarios: list, style: str = "section") -> str:
    """整篇markdown。

    一张表里有多个故障场景时，每个场景一个 `## 场景N：…` 小节、步骤降为 `###`，
    否则场景边界会丢掉——第二个场景的步骤2、步骤3 会和第一个场景的步骤平铺在
    同一层，蒸馏时两个场景就混成一条排查链了。只有一个无名场景时退回原来的
    "整篇一串步骤"形态。
    """
    branch = style == "branch"
    lines = [f"# {title}", ""]
    named = [s for s in scenarios if s["title"]]

    if not named:
        steps = [step for s in scenarios for step in s["steps"]]
        labels = step_labels(steps)
        if branch:
            lines += ["## 排障步骤", ""] + render_steps_branch(steps, labels)
        else:
            if len(steps) > 1:
                lines += ["## 排障步骤总览", ""]
                lines += [f"{seq}. {label}" for seq, label in enumerate(labels, 1)]
                lines.append("")
            lines += render_steps(steps, labels, "##")
        return re.sub(r"\n{3,}", "\n\n", "\n".join(lines)).strip() + "\n"

    all_labels = [step_labels(s["steps"]) for s in scenarios]

    lines += ["## 排障场景总览", ""]
    for seq, (scenario, labels) in enumerate(zip(scenarios, all_labels), 1):
        lines.append(f"{seq}. 场景{seq}：{scenario['title'] or '（未命名场景）'}")
        lines += [f"    1. {label}" if i == 0 else f"    {i + 1}. {label}"
                  for i, label in enumerate(labels)]
    lines.append("")

    for seq, (scenario, labels) in enumerate(zip(scenarios, all_labels), 1):
        lines += [f"## 场景{seq}：{scenario['title'] or '（未命名场景）'}", ""]
        # 排障目标那一格常带着告警名和故障构造方法，只有一行时标题已经写过了
        if scenario["goal"] and "\n" in scenario["goal"]:
            lines += ["**排障目标**", "", scenario["goal"], ""]
        if scenario["topology"]:
            lines += ["**组网场景**", "", scenario["topology"], ""]
        lines += (render_steps_branch(scenario["steps"], labels) if branch
                  else render_steps(scenario["steps"], labels, "###"))

    return re.sub(r"\n{3,}", "\n\n", "\n".join(lines)).strip() + "\n"


def read_workbook(path: str) -> dict:
    """读入表格，返回 {sheet名: [[单元格]]}。"""
    ext = os.path.splitext(path)[1].lower()
    if ext in (".csv", ".tsv", ".txt"):
        # 从Excel直接复制粘贴出来的就是制表符分隔、多行字段带引号，csv模块能正确还原
        delimiter = "\t" if ext in (".tsv", ".txt") else ","
        with open(path, "r", encoding="utf-8-sig", newline="") as f:
            rows = [row for row in csv.reader(f, delimiter=delimiter)]
        return {os.path.splitext(os.path.basename(path))[0]: rows}

    if ext not in (".xlsx", ".xlsm"):
        raise ValueError(f"不支持的文件类型: {ext}（支持 .xlsx/.xlsm/.csv/.tsv）")
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise ValueError(
            "读取 .xlsx 需要 openpyxl，请先 pip install openpyxl；"
            "或把表格另存/粘贴成 .tsv 再转（.tsv 不需要额外依赖）")
    # data_only=True 取公式的计算结果而不是公式本身
    workbook = load_workbook(path, data_only=True, read_only=True)
    sheets = {}
    for sheet in workbook.worksheets:
        sheets[sheet.title] = [list(row) for row in sheet.iter_rows(values_only=True)]
    workbook.close()
    return sheets


def safe_filename(name: str) -> str:
    return FILENAME_UNSAFE.sub("_", name).strip() or "sheet"


def convert_sheet(name: str, rows: list, title: str, skip_unsupported: bool,
                  style: str = "section") -> tuple:
    """返回 (markdown文本, 统计信息)。"""
    scenarios, extras, header_idx = parse_rows(rows)

    dropped = []
    if skip_unsupported:
        for scenario in scenarios:
            kept = []
            for step in scenario["steps"]:
                if "不支持" in step.get("version", ""):
                    title_line = (step.get("step_title") or "").split("\n")[0].strip()
                    dropped.append(f"步骤{step.get('step_no', '?')} {title_line}")
                else:
                    kept.append(step)
            scenario["steps"] = kept
        scenarios = [s for s in scenarios if s["steps"]]

    step_count = sum(len(s["steps"]) for s in scenarios)
    if not step_count:
        raise ValueError("表头之下没有解析出任何步骤行")

    # "修复建议"空着、"影响性"里却写着修复动作，多半是这一行整体右移了一列。
    # 本脚本严格按表头列位映射，源表错位就会照着错位输出，所以点出来让人回表里核对。
    shifted = [f"步骤{step.get('step_no', '?')}"
               for scenario in scenarios for step in scenario["steps"]
               if not step.get("fix") and step.get("impact")]

    text = render_document(title or name, scenarios, style)
    links = sorted(set(LINK_PATTERN.findall(text)))
    return text, {"steps": step_count, "header_row": header_idx + 1,
                  "scenarios": [s["title"] for s in scenarios if s["title"]],
                  "extras": [n for _, n in extras], "dropped": dropped,
                  "links": links, "shifted": shifted}


def main(source: str, out_dir: str, sheet_filter: str, title: str,
         tree: str, skip_unsupported: bool, to_stdout: bool,
         style: str = "section"):
    print("=" * 60)
    print("Excel排障步骤表 → markdown源文档")
    print("=" * 60)
    print(f"输入: {source}")

    if not os.path.isfile(source):
        print(f"错误: 文件不存在: {source}")
        sys.exit(1)
    try:
        sheets = read_workbook(source)
    except ValueError as e:
        print(f"错误: {e}")
        sys.exit(1)

    if sheet_filter:
        sheets = {k: v for k, v in sheets.items() if sheet_filter in k}
        if not sheets:
            print(f"错误: 没有匹配 --sheet {sheet_filter!r} 的工作表")
            sys.exit(1)

    # --tree 直接摆成蒸馏源树的布局：<一级目录>/<二级目录>/<sheet>.md，
    # 一个二级目录蒸馏出一篇skill（见 skill_self_distill_pipeline 的分组规则）
    target_dir = os.path.join(out_dir, *tree.split("/")) if tree else out_dir
    if not to_stdout:
        print(f"输出目录: {target_dir}")
    print()

    failed = 0
    for name, rows in sheets.items():
        try:
            text, stats = convert_sheet(name, rows, title if len(sheets) == 1 else "",
                                        skip_unsupported, style)
        except ValueError as e:
            print(f"[FAIL] 工作表 {name!r}: {e}")
            failed += 1
            continue

        if to_stdout:
            print(f"{'=' * 20} {name} {'=' * 20}\n{text}")
        else:
            os.makedirs(target_dir, exist_ok=True)
            path = os.path.join(target_dir, safe_filename(name) + ".md")
            with open(path, "w", encoding="utf-8") as f:
                f.write(text)
            print(f"[OK] {name} → {path}"
                  f"（{stats['steps']} 步，表头在第{stats['header_row']}行，{len(text)}字符）")

        if stats["scenarios"]:
            # 一张表里有多个故障场景是常态，列出来核对场景边界切对了没有
            print(f"     {len(stats['scenarios'])} 个故障场景: "
                  f"{'、'.join(stats['scenarios'])}")
        if stats["shifted"]:
            print(f"     [WARN] {'、'.join(stats['shifted'])} 的\"修复建议\"为空、"
                  f"\"影响性\"却有内容，疑似源表这一行整列右移，请回表核对")
        if stats["extras"]:
            print(f"     未识别的列已原样保留为'其它信息': {'、'.join(stats['extras'])}")
        if stats["dropped"]:
            print(f"     按 --skip-unsupported 丢弃 {len(stats['dropped'])} 行")
        if stats["links"]:
            # 蒸馏的写作约束要求删掉链接（转换后已失效），这里先报出来，
            # 免得它们一路带进skill
            print(f"     [WARN] 正文含 {len(stats['links'])} 条外部链接，蒸馏时会被要求删除，"
                  f"必要的内容请先摘成文字：")
            for link in stats["links"]:
                print(f"       - {link}")

    print()
    if failed:
        print(f"共 {len(sheets)} 张表，失败 {failed} 张")
        sys.exit(1)
    print(f"共转换 {len(sheets)} 张表。")
    if not to_stdout:
        print("下一步：确认内容后，把目录摆成 <源树>/<一级目录>/<二级目录>/*.md "
              "（一个二级目录=一篇skill），再跑 skill_self_distill_pipeline.py。")


def _get_option(args: list, flag: str) -> str:
    if flag not in args:
        return ""
    idx = args.index(flag)
    if idx + 1 >= len(args):
        print(f"错误: {flag} 后面缺少取值")
        sys.exit(1)
    value = args[idx + 1]
    del args[idx:idx + 2]
    return value


if __name__ == "__main__":
    argv = sys.argv[1:]
    skip_unsupported = "--skip-unsupported" in argv
    to_stdout = "--stdout" in argv
    argv = [a for a in argv if a not in ("--skip-unsupported", "--stdout")]
    sheet_filter = _get_option(argv, "--sheet")
    doc_title = _get_option(argv, "--title")
    tree_path = _get_option(argv, "--tree")
    doc_style = _get_option(argv, "--style") or "section"

    if doc_style not in ("section", "branch"):
        print(f"错误: --style 只支持 section / branch，收到 {doc_style!r}")
        sys.exit(1)
    if not argv:
        print(__doc__)
        sys.exit(1)
    default_out = os.path.join(
        "docs_from_excel", os.path.splitext(os.path.basename(argv[0]))[0])
    main(argv[0], argv[1] if len(argv) > 1 else default_out,
         sheet_filter, doc_title, tree_path, skip_unsupported, to_stdout, doc_style)
